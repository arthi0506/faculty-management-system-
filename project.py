#this is the python program to create the faculty free slot management system using TKinter.
#*Here a gui application is created where user can search faculty free slots by their id or name
#*also user can see brach wise timetable
#*and even faculty can enter their free slots in excel sheet 
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
from PIL import Image, ImageTk
from tkinter import PhotoImage

#creating the gui interface 
root = tk.Tk()
root .geometry('800x500') #setting the size
root.title('faculty free slot manger') #naming the window

#creating the search page to search faculty details
def search_page():
    search_frame=tk.Frame(main_frame)
    show_frame=tk.Frame(main_frame,highlightbackground='black',
                        highlightthickness=2)
    #this function is for search page button 
    def click(event):
        entry2.config(state='normal')
        entry2.delete(0,'end')

    #this func is for clear button in the search page
    def clear():
        entry1.delete(0,'end')
        entry2.delete(0,'end')
        ename.delete(0,'end')
        esub.delete(0,'end')
        emon.delete(0,'end')
        etues.delete(0,'end')
        ewed.delete(0,'end')
        ethur.delete(0,'end')
        efri.delete(0,'end')
        esat.delete(0,'end')
        result = tk.Label(search_frame, text="                                       ", font=('Bold', 10), fg='red')
        result.grid(row=5, column=0, columnspan=3)

    #sub clear is function called internally by serch button to clear the perviously searched data     
    def subclear():
        ename.delete(0,'end')
        esub.delete(0,'end')
        emon.delete(0,'end')
        etues.delete(0,'end')
        ewed.delete(0,'end')
        ethur.delete(0,'end')
        efri.delete(0,'end')
        esat.delete(0,'end')
        result = tk.Label(search_frame, text="                                      ", font=('Bold', 10), fg='red')
        result.grid(row=5, column=0, columnspan=3)
    
    #this func is for the search button in search page
    def show():
        try:
            subclear()
            excel_path = r'C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/faculty_proj.xlsx'
            search_id = entry1.get()
            NAME=entry2.get()
            faculty_details = None
            file = openpyxl.load_workbook(excel_path)
            sheet = file.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0])==search_id or row[1]==NAME:
                        faculty_details = row
                        break
            file.close    
            if faculty_details:
                    result = tk.Label(search_frame, text="DETAILS FOUND !!", font=('Bold', 10), fg='green')
                    result.grid(row=5, column=0, columnspan=3)
                    ename.insert(0,row[1])
                    esub.insert(0,row[2])
                    emon.insert(0,row[3])
                    etues.insert(0,row[4])
                    ewed.insert(0,row[5])
                    ethur.insert(0,row[5])
                    efri.insert(0,row[5])
                    esat.insert(0, row[8])
            else:
                    result = tk.Label(search_frame, text="      DETAILS NOT FOUND!!      ", font=('Bold', 10), fg='red')
                    result.grid(row=5, column=0, columnspan=3)
        except Exception as e:
           print(str(e))
                    


    #creating the widgets for search frame
    lb0=tk.Label(search_frame,text='ENTER FACULTY ID,NAME TO SEE THEIR FREE SLOTS OF A WEEK BELOW ',
                 font=('Bold',18),bg='old lace',wraplength=630)
    lb1=tk.Label(search_frame,text='faculty\n name :',font=('Bold',18))
    lb2=tk.Label(search_frame,text=' id no. :',font=('Bold',18))
    entry1=tk.Entry(search_frame,width=30,font='timesNewRoman')
    entry2=tk.Entry(search_frame,width=30,font='timesNewRoman')
    entry2.insert(0,"(optional)")
    entry2.config(state="disable")
    entry2.bind("<Button-1>",click) #button-1 means when you right click on widget it enables it
    show=tk.Button(search_frame,text='show free\nslots',font=('Bold',15),fg='#158aff',command=show)
    cl=tk.Button(search_frame,text='clear',font=('Bold',18),command=lambda:clear())

    lb0.grid(row=0,column=0,columnspan=3,padx=10,pady=10)
    lb1.grid(row=3,column=0,padx=10,pady=10)
    lb2.grid(row=2,column=0,padx=5,pady=5)
    entry1.grid(row=2,column=1,padx=5,pady=5)
    entry2.grid(row=3,column=1,padx=5,pady=5)
    show.grid(row=4,column=1,columnspan=2,padx=5,pady=5)
    cl.grid(row=4,column=0,padx=5,pady=5)

    #cearting widgets for show page to show searched data
    name_lable=tk.Label(show_frame,text='       NAME :       ',font=('Bold',15),bg='old lace',wraplength=250)
    sub_lable=tk.Label(show_frame,text='     SUBJECT :     ',font=('Bold',15),bg='old lace',wraplength=250)
    mon=tk.Label(show_frame,text='MONDAY',font=('Bold',15))
    tues=tk.Label(show_frame,text='TUESDAY',font=('Bold',15))
    wed=tk.Label(show_frame,text='WEDNESDAY',font=('Bold',15))
    thur=tk.Label(show_frame,text='THURSDAY',font=('Bold',15))
    fri=tk.Label(show_frame,text='FRIDAY',font=('Bold',15))
    sat=tk.Label(show_frame,text='SATURDAY',font=('Bold',15))
    ename=tk.Entry(show_frame,width=20,font='timesNewRoman',fg="green")
    esub=tk.Entry(show_frame,width=20,font='timesNewRoman',fg="green")
    emon=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")
    etues=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")
    ewed=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")
    ethur=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")
    efri=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")
    esat=tk.Entry(show_frame,width=17,font='timesNewRoman',justify="center",fg="green")

    name_lable.grid(row=0,column=0)
    ename.grid(row=0,column=1)
    sub_lable.grid(row=1,column=0)
    esub.grid(row=1,column=1)
    mon.grid(row=3,column=0)
    tues.grid(row=3,column=1)
    wed.grid(row=3,column=2)
    thur.grid(row=5,column=0)
    fri.grid(row=5,column=1)
    sat.grid(row=5,column=2)
    emon.grid(row=4,column=0)
    etues.grid(row=4,column=1)
    ewed.grid(row=4,column=2)
    ethur.grid(row=6,column=0)
    efri.grid(row=6,column=1)
    esat.grid(row=6,column=2)

    #placing search_frame(in this page we can search by faculty id or faculty name)
    search_frame.pack(side=tk.TOP)
    search_frame.pack_propagate(True)
    search_frame.configure(width=650,height=250)


    #placing show_frame (in this page free slots will been displayed)
    show_frame.pack(side=tk.BOTTOM)
    show_frame.pack_propagate(True)
    show_frame.configure(width=650,height=250)

#this func is used to create the table page 
def table_page():
    table_frame=tk.Frame(main_frame)
    #this one is for displaying the image and it is calle internally when any button is clicked
    def display_image(image_path):
            delete_pages()
            page = tk.Frame(main_frame)
            page.pack(pady=20)
            image_label = tk.Label(page, image=None)
            image_label.pack() 
            img = Image.open(image_path)
            img = img.resize((500, 400), Image.LANCZOS)
            img = ImageTk.PhotoImage(img)
            image_label.config(image=img)
            image_label.image = img
    #thid functions are for each button where image location is given
    def select_csit():
        display_image("C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/py_project_img/csit_pro.jpeg")
    def select_cse():
        display_image("C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/py_project_img/cse_pro.jpeg")
    def select_csm():
        display_image("C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/py_project_img/cse_pro.jpeg")
    def select_csd():
        display_image("C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/py_project_img/csd_pro.jpeg")
    def select_it():
        display_image("C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/py_project_img/it_Pro.jpeg")

    #creating the widgets for table page
    lb=tk.Label(table_frame,text='      To see time table Select the branch :     ',font=('Bold',30),bg='old lace',wraplength=630)
    cse=tk.Button(table_frame,text='CSE',font=('Bold',18),bg='white',command=lambda:select_cse())
    csm=tk.Button(table_frame,text='CSM',font=('Bold',18),bg='white',command=lambda:select_csm())
    csd=tk.Button(table_frame,text='CSD',font=('Bold',18),bg='white',command=lambda:select_csd())
    csit=tk.Button(table_frame,text='CSIT',font=('Bold',18),bg='white',command=lambda:select_csit())
    it=tk.Button(table_frame,text='IT',font=('Bold',18),bg='white',command=lambda:select_it())
    #placing the widgets
    lb.grid(row=0,column=0)
    csit.grid(row=2,column=0,padx=10,pady=5)
    it.grid(row=4,column=0,padx=10,pady=5)
    cse.grid(row=6,column=0,padx=10,pady=5)
    csm.grid(row=8,column=0,padx=10,pady=5)
    csd.grid(row=10,column=0,padx=10,pady=5)

    table_frame.pack(pady=20) #setting the page 

#from here functions are defined for entry page where faculty can enter data to excel
def entry_page():
    entry_frame=tk.Frame(main_frame)
    def clear():
        name_entry.delete(0,'end')
        id_entry.delete(0,'end')
        sub_entry.delete(0,'end')
        mon_entry.delete(0,'end')
        tues_entry.delete(0,'end')
        wed_entry.delete(0,'end')
        thu_entry.delete(0,'end')
        fri_entry.delete(0,'end')
        sat_entry.delete(0,'end')
        result = tk.Label(entry_frame, text="                             ", font=('Bold', 10), fg='red')
        result.grid(row=9, column=0, columnspan=3)
    def entry():
        # Get the data from the GUI
        name = name_entry.get()
        id = id_entry.get()
        sub = sub_entry.get()
        monday=mon_entry.get()
        tuesday=tues_entry.get()
        wednesday=wed_entry.get()
        thursday=thu_entry.get()
        friday=fri_entry.get()
        saturday=sat_entry.get()

        # Load the Excel file
        workbook = openpyxl.load_workbook('C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/faculty_proj.xlsx')

        # Select the active worksheet
        worksheet = workbook.active

        # Find the next available row in column A
        row = worksheet.max_row + 1

        # Write the data to the worksheet
        worksheet.cell(row=row, column=1, value=id)
        worksheet.cell(row=row, column=2, value=name)
        worksheet.cell(row=row, column=3, value=sub)
        worksheet.cell(row=row, column=4, value=monday)
        worksheet.cell(row=row, column=5, value=tuesday)
        worksheet.cell(row=row, column=6, value=wednesday)
        worksheet.cell(row=row, column=7, value=thursday)
        worksheet.cell(row=row, column=8, value=friday)
        worksheet.cell(row=row, column=9, value=saturday)

        # Save the changes to the Excel file
        workbook.save('C:/Users/Arthi kumari singh/OneDrive/Desktop/py project/faculty_proj.xlsx')

        # Display a message to the user
        result = tk.Label(entry_frame, text="enterd successfully!!",fg='green', font=('Bold', 10))
        result.grid(row=9, column=0, columnspan=3)
        workbook.close

    #creating the widgets for entry page
    lb=tk.Label(entry_frame,text='     only for faculty       ',font=('Bold',30),bg='old lace',wraplength=630)
    f_name=tk.Label(entry_frame,text='Faculty name:',font=25)
    f_id=tk.Label(entry_frame,text='Faculty id no.:',font=25)
    mon=tk.Label(entry_frame,text='Monday:',font=25)
    tues=tk.Label(entry_frame,text='Tuesday:',font=25)
    wed=tk.Label(entry_frame,text='Wednesday:',font=25)
    thu=tk.Label(entry_frame,text='Thursday:',font=25)
    fri=tk.Label(entry_frame,text='Friday:',font=25)
    sat=tk.Label(entry_frame,text='Saturday:',font=25)
    slot=tk.Label(entry_frame,text='Enter your free time with Room number below:',font=25)
    sub=tk.Label(entry_frame,text='subject:',font=25)
    name_entry=tk.Entry(entry_frame,width=40,font='timesNewRoman')
    id_entry=tk.Entry(entry_frame,width=40,font='timesNewRoman')
    mon_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    tues_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    wed_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    thu_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    fri_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    sat_entry=tk.Entry(entry_frame,width=15,font='timesNewRoman')
    sub_entry=tk.Entry(entry_frame,width=30,font='timesNewRoman')
    clear=tk.Button(entry_frame,text='clear',bg='white',font=30,command=clear)
    entry=tk.Button(entry_frame,text='Enter',bg='white',fg='blue',font=30,command=entry)

    #placing the widgets
    lb.grid(row=0,column=0,columnspan=3,padx=10,pady=10)
    f_id.grid(row=1,column=0,padx=10,pady=10)
    id_entry.grid(row=1,column=1,columnspan=4,padx=0,pady=0)
    f_name.grid(row=2,column=0,padx=10,pady=10)
    name_entry.grid(row=2,column=1,columnspan=4,padx=0,pady=0)
    slot.grid(row=3,column=0,columnspan=3,padx=10,pady=10)
    mon.grid(row=4,column=0,padx=10,pady=10)
    mon_entry.grid(row=4,column=1,padx=0,pady=0)
    tues.grid(row=4,column=2,padx=10,pady=10)
    tues_entry.grid(row=4,column=3,padx=0,pady=0)
    wed.grid(row=5,column=0,padx=10,pady=10)
    wed_entry.grid(row=5,column=1,padx=10,pady=10)
    thu.grid(row=5,column=2,padx=10,pady=10)
    thu_entry.grid(row=5,column=3,padx=10,pady=10)
    fri.grid(row=6,column=0,padx=10,pady=10)
    fri_entry.grid(row=6,column=1,padx=10,pady=10)
    sat.grid(row=6,column=2,padx=10,pady=10)
    sat_entry.grid(row=6,column=3,padx=10,pady=10)
    sub.grid(row=7,column=0,padx=5,pady=5)
    sub_entry.grid(row=7,column=1,columnspan=2,padx=5,pady=5)
    clear.grid(row=9,column=0,padx=15,pady=15)
    entry.grid(row=9,column=2,padx=10,pady=10)

    entry_frame.pack(pady=20)

#this funtion is for hiding the indicator
def hide_indicators():
    search_indicate.config(bg='#c3c3c3')
    table_indicate.config(bg='#c3c3c3')
    entry_indicate.config(bg='#c3c3c3')

#for deleting the pervious open page
def delete_pages():
    for frame in main_frame.winfo_children():
        frame.destroy()    

#for displaying the indicator
def indicate(lb, page):
    hide_indicators()
    lb.config(bg='#158aff')
    delete_pages()
    page()

#for getting back to home page when home label(internally button) is cliked
def home():
    try:
        home_frame=tk.Frame(main_frame)
        delete_pages()
        main_lb=tk.Label(main_frame,text=' welcome...! to faculty free slot management system ',
                    font=('time',35,"italic"),fg='black',bd=0,
                    bg='old lace',wraplength=500)
        main_lb.place(x=70,y=130)

        home_frame.pack(pady=20)
    except Exception as e :
         hide_indicators()

#creating left side of the page
option_frame=tk.Frame(root,bg='#c3c3c3')

option_frame.pack(side=tk.LEFT)
option_frame.pack_propagate(False)
option_frame.configure(width=150,height=500)

#creating and placing widgets for left side of page
#home page is the first page 
home_lable=tk.Button(option_frame,text='☰ HOME ',font=('bold',20),
                bd=0,bg='#c3c3c3', fg='black',command=lambda:home())
home_lable.place(x=8,y=50)

#creating search page
search_btn=tk.Button(option_frame,text='search',font=('bold',20),
                     fg='#158aff',bd=0,bg='white',
                   command=lambda:indicate(search_indicate, search_page))
search_btn.place(x=15,y=120)
#creating indicator for search page
search_indicate=tk.Label(option_frame,text='',bg='#c3c3c3')
search_indicate.place(x=3,y=120,width=5,height=40)


#creating page where to see all time table
table_btn=tk.Button(option_frame,text='see time\n table',font=('bold',20),
                    fg='#158aff',bd=0,bg='white',
                   command=lambda:indicate(table_indicate, table_page))
table_btn.place(x=15,y=200)

table_indicate=tk.Label(option_frame,text='',bg='#c3c3c3') #creating indicator 
table_indicate.place(x=3,y=215,width=5,height=50)

#creating entry page to enter faculty data in database
entry_btn=tk.Button(option_frame,text='faculty \n edit slot',font=('bold',20),
                    fg='#158aff',bd=0,bg='white',
                   command=lambda:indicate(entry_indicate, entry_page))
entry_btn.place(x=15,y=310)

entry_indicate=tk.Label(option_frame,text='',bg='#c3c3c3') #creating indicator
entry_indicate.place(x=3,y=320,width=5,height=42)

main_frame=tk.Frame(root,highlightbackground='black',highlightthickness=2)

main_lb=tk.Label(main_frame,text=' welcome...! to faculty free slot management system ',
                 font=('time',35,"italic"),fg='black',bd=0,
                 bg='old lace',wraplength=500)
main_lb.place(x=70,y=130)
main_frame.pack(side=tk.LEFT)
main_frame.pack_propagate(False)
main_frame.configure(width=650,height=500)

root.mainloop()